"""Parse prior-year CoreTax SPT XLSX into the staging table.

Reads a filled-in template (the prior-year submission) and extracts:
  - C1/C2/C3 for taxpayer metadata
  - Row 4 headers to identify column roles (E=prior_tax_year, F=target_tax_year, G=market, H=keterangan)
  - Rows 6 → first blank or row 47 for asset/liability data

Writes raw cell text + parsed normalized values into coretax_import_staging.
"""
from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any

import openpyxl

from finance.coretax.db import (
    _norm,
    _slug,
    _utcnow,
    ensure_coretax_tables,
    get_asset_codes,
)


class ImportParseError(ValueError):
    pass


def _parse_idr_value(cell_value) -> float | None:
    """Parse a cell that may contain an IDR-formatted number or a plain float."""
    if cell_value is None:
        return None
    text = str(cell_value).strip()
    if not text or text == "-":
        return None
    # Remove thousands separators (Indonesian: 1.234.567,89) and currency symbols
    text = text.replace("Rp", "").replace("rp", "").strip()
    # If it contains both dots and commas, dots are thousands separators
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "." in text:
        # Ambiguous: could be thousands (1.234.567) or decimal (1234.567)
        # For IDR-scale values, dots are almost always thousands separators
        parts = text.split(".")
        if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3):
            text = text.replace(".", "")
    text = text.replace(",", "")
    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def _detect_tax_years_from_headers(ws, header_row: int) -> tuple[int, int]:
    """Detect prior_tax_year and target_tax_year from E and F header cells.

    E header typically says something like "2024" and F header says "2025".
    Returns (prior_tax_year, target_tax_year).
    """
    e_text = str(ws[f"E{header_row}"].value or "").strip()
    f_text = str(ws[f"F{header_row}"].value or "").strip()

    e_year = _extract_year(e_text)
    f_year = _extract_year(f_text)

    if e_year and f_year:
        return e_year, f_year
    # Fallback: if we only get one, infer the other
    if e_year:
        return e_year, e_year + 1
    if f_year:
        return f_year - 1, f_year
    raise ImportParseError("Cannot determine tax years from E/F column headers")


def _extract_year(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text)
    return int(match.group(1)) if match else None


def _heuristic_stable_key(kode_harta: str, keterangan: str,
                           acquisition_year: int | None) -> str:
    """Try to generate a PWM-style stable key from parsed row content.

    Falls back to manual:... if no PWM match signature detected.
    """
    from finance.coretax.db import make_stable_key_manual
    return make_stable_key_manual(kode_harta, keterangan, acquisition_year)


def parse_prior_year_xlsx(
    file_path: Path,
    target_tax_year: int,
    conn,
) -> dict:
    """Parse a prior-year XLSX template and write rows to staging.

    Args:
        file_path: Path to the XLSX file.
        target_tax_year: The SPT year being prepared (e.g. 2025).
        conn: SQLite connection (from open_db).

    Returns:
        dict with batch_id, row_count, warnings, prior_tax_year.
    """
    ensure_coretax_tables(conn)

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    if not wb.worksheets:
        raise ImportParseError("Workbook has no sheets")
    ws = wb.worksheets[0]

    # ── Extract taxpayer metadata from C1, C2, C3 ────────────────────────
    nama_wp = str(ws["C1"].value or "").strip()
    npwp = str(ws["C2"].value or "").strip()
    # C3 often has notes or additional taxpayer info

    # ── Find header row (B == "Kode Harta") ──────────────────────────────
    header_row = None
    for row_idx in range(1, 11):
        if str(ws[f"B{row_idx}"].value or "").strip() == "Kode Harta":
            header_row = row_idx
            break
    if header_row is None:
        raise ImportParseError("Template missing 'Kode Harta' header in B1:B10")

    # ── Detect tax years from headers ────────────────────────────────────
    prior_tax_year, detected_target = _detect_tax_years_from_headers(ws, header_row)
    if target_tax_year not in (prior_tax_year, detected_target):
        raise ImportParseError(
            f"Template year mismatch: uploaded template has columns for "
            f"{prior_tax_year} (E) and {detected_target} (F); "
            f"selected year {target_tax_year} is not in this template."
        )
    # Determine which column is the carry-forward source for this year
    use_e_as_carry = (target_tax_year == prior_tax_year)

    # ── Load asset codes for carry-forward rules ─────────────────────────
    asset_codes = {ac["kode"]: ac for ac in get_asset_codes(conn)}

    # ── Generate batch ID and wipe any existing staging for this year ────
    batch_id = uuid.uuid4().hex[:12]
    conn.execute(
        "DELETE FROM coretax_import_staging WHERE target_tax_year = ?",
        (target_tax_year,),
    )

    # ── Parse data rows ──────────────────────────────────────────────────
    first_data_row = header_row + 1
    warnings: list[str] = []
    row_count = 0
    now = _utcnow()

    # C column layout in the template:
    # The template uses column C for the sequential number (row numbering),
    # column H for the description/keterangan (long text). C may contain a short
    # label but H is the authoritative description for matching.
    # Template structure:
    #   A = No, B = Kode Harta, C = sequential number / short label, D = Tahun Perolehan
    #   E = prior year value, F = current year value, G = market value, H = Keterangan (long)
    # parsed_keterangan reads H first, falls back to C.

    row_idx = first_data_row
    while row_idx <= 47:  # Template rows 6-47 = 42 data rows max
        # Check for terminator
        col_c_val = str(ws[f"C{row_idx}"].value or "").strip()
        col_h_val = str(ws[f"H{row_idx}"].value or "").strip()
        if col_c_val.upper() == "TOTAL ASET KOTOR":
            break
        # Also check if the row is completely empty (no kode, no description)
        col_b_val = str(ws[f"B{row_idx}"].value or "").strip()
        if not col_b_val and not col_c_val and not col_h_val:
            # Skip completely blank rows but keep going (template may have gaps)
            row_idx += 1
            continue

        # Raw cell values as text for audit trail
        source_col_b = col_b_val
        source_col_c = col_c_val
        source_col_d = str(ws[f"D{row_idx}"].value or "").strip()
        source_col_e = str(ws[f"E{row_idx}"].value or "").strip()
        source_col_f = str(ws[f"F{row_idx}"].value or "").strip()
        source_col_g = str(ws[f"G{row_idx}"].value or "").strip()
        source_col_h = col_h_val

        # Parsed values
        parsed_kode = source_col_b if source_col_b else None
        parsed_keterangan = source_col_h or source_col_c or ""
        parsed_acq_year = None
        if source_col_d:
            try:
                parsed_acq_year = int(float(source_col_d))
            except (ValueError, TypeError):
                pass

        parsed_e = _parse_idr_value(ws[f"E{row_idx}"].value)
        parsed_f = _parse_idr_value(ws[f"F{row_idx}"].value)
        parsed_g = _parse_idr_value(ws[f"G{row_idx}"].value)

        parse_warning = None
        if parsed_e is None and source_col_e:
            parse_warning = f"Could not parse E value: {source_col_e!r}"
        if parsed_f is None and source_col_f:
            parse_warning = (parse_warning + "; " if parse_warning else "") + f"Could not parse F value: {source_col_f!r}"

        # Determine kind (asset or liability)
        # Liabilities typically appear in a separate section with different kode patterns
        # For now, default to 'asset' unless the keterangan indicates liability
        parsed_kind = "asset"
        if parsed_kode and parsed_kode.startswith("L"):
            parsed_kind = "liability"

        # Carry-forward rule from asset codes
        rule_carry = None
        if parsed_kode and parsed_kode in asset_codes:
            rule_carry = asset_codes[parsed_kode]["default_carry_forward"]

        # Proposed stable key
        proposed_key = _heuristic_stable_key(parsed_kode or "", parsed_keterangan, parsed_acq_year)

        parsed_carry = parsed_e if use_e_as_carry else parsed_f
        parsed_prior = parsed_f if use_e_as_carry else parsed_e

        conn.execute(
            """INSERT INTO coretax_import_staging
               (staging_batch_id, target_tax_year, source_file_name, source_sheet_name,
                source_row_no, source_col_b_kode, source_col_c_keterangan,
                source_col_d_acq_year, source_col_e_value, source_col_f_value,
                source_col_g_value, source_col_h_note,
                parsed_kode_harta, parsed_keterangan, parsed_acquisition_year,
                parsed_prior_amount_idr, parsed_carry_amount_idr, parsed_market_value_idr,
                parsed_kind, proposed_stable_key, rule_default_carry_forward,
                user_override_carry_forward, parse_warning, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)""",
            (batch_id, target_tax_year, file_path.name, ws.title or "Sheet1",
             row_idx, source_col_b, source_col_c,
             source_col_d, source_col_e, source_col_f,
             source_col_g, source_col_h,
             parsed_kode, parsed_keterangan, parsed_acq_year,
             parsed_prior, parsed_carry, parsed_g,
             parsed_kind, proposed_key, rule_carry,
             parse_warning, now),
        )
        row_count += 1
        row_idx += 1

    # ── Store taxpayer metadata ──────────────────────────────────────────
    from finance.coretax.db import upsert_taxpayer
    upsert_taxpayer(conn, target_tax_year, nama_wp, npwp)

    conn.commit()

    return {
        "batch_id": batch_id,
        "row_count": row_count,
        "prior_tax_year": prior_tax_year,
        "warnings": warnings,
    }
