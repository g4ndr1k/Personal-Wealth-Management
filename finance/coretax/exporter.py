"""CoreTax XLSX exporter — writes coretax_rows into a canonical template.

Loads the template from data/coretax/templates/, fills cells for the given
tax_year, and saves versioned output to data/coretax/output/.

Key rules:
  - F-cell: if current_amount_source == 'carried_forward' → write formula '=E{row}'
            otherwise write literal current_amount_idr (or leave blank if NULL)
  - Template capacity: 42 data rows (6-47). Raise if exceeded.
  - Does not touch rows 48+ formulas/styles.
  - Writes audit JSON sidecar.
"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl

from finance.coretax.db import (
    _utcnow,
    ensure_coretax_tables,
    get_rows_for_year,
    get_taxpayer,
)


class ExportError(ValueError):
    pass


MAX_DATA_ROWS = 42  # Template rows 6-47
FIRST_DATA_ROW = 6
LAST_DATA_ROW = 47


@dataclass
class ExportAuditRow:
    xlsx_row: int
    stable_key: str
    kode_harta: str
    keterangan: str
    prior_amount_idr: float | None
    current_amount_idr: float | None
    market_value_idr: float | None
    current_amount_source: str | None
    f_cell_rule: str  # 'formula' or 'literal' or 'blank'


@dataclass
class ExportResult:
    tax_year: int
    file_id: str
    total_rows: int
    total_prior: float
    total_current: float
    total_market: float
    liability_total: float
    rows: list[ExportAuditRow]


def export_coretax_xlsx(
    conn,
    tax_year: int,
    template_dir: str | Path,
    output_dir: str | Path,
) -> ExportResult:
    """Export coretax_rows for the given tax_year into an XLSX file.

    Returns ExportResult with file_id and audit data.
    Raises ExportError if template not found or row count exceeds capacity.
    """
    ensure_coretax_tables(conn)

    template_dir = Path(template_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Load rows ────────────────────────────────────────────────────────
    rows = get_rows_for_year(conn, tax_year)
    if not rows:
        raise ExportError(f"No coretax_rows found for tax_year {tax_year}")

    # Separate assets and liabilities
    assets = [r for r in rows if r["kind"] == "asset"]
    liabilities = [r for r in rows if r["kind"] == "liability"]

    # Assets go into template rows 6-47 (42 max)
    if len(assets) > MAX_DATA_ROWS:
        raise ExportError(
            f"Template capacity exceeded: {len(assets)} asset rows but only "
            f"{MAX_DATA_ROWS} slots available (rows {FIRST_DATA_ROW}-{LAST_DATA_ROW}). "
            f"Reduce rows or expand template."
        )

    # ── Find template ────────────────────────────────────────────────────
    template_path = _find_template(template_dir)
    if not template_path:
        raise ExportError(
            f"No XLSX template found in {template_dir}. "
            "Copy the yearly CoreTax workbook into that directory."
        )

    # ── Load taxpayer info ───────────────────────────────────────────────
    taxpayer = get_taxpayer(conn, tax_year)

    # ── Open template workbook ───────────────────────────────────────────
    wb = openpyxl.load_workbook(str(template_path), keep_vba=False)
    if not wb.worksheets:
        raise ExportError("Template workbook has no sheets")
    ws = wb.worksheets[0]

    # ── Write taxpayer metadata (C1, C2, C3) ────────────────────────────
    if taxpayer:
        if taxpayer.get("nama_wajib_pajak"):
            ws["C1"] = taxpayer["nama_wajib_pajak"]
        if taxpayer.get("npwp"):
            ws["C2"] = taxpayer["npwp"]

    # ── Write asset rows ─────────────────────────────────────────────────
    audit_rows: list[ExportAuditRow] = []
    for idx, row in enumerate(assets):
        xlsx_row = FIRST_DATA_ROW + idx
        _write_data_row(ws, xlsx_row, row, audit_rows)

    # Clear any remaining template rows that weren't filled
    for idx in range(len(assets), MAX_DATA_ROWS):
        xlsx_row = FIRST_DATA_ROW + idx
        _clear_data_row(ws, xlsx_row)

    # ── Save workbook ────────────────────────────────────────────────────
    file_id = _next_output_filename(output_dir, tax_year)
    output_path = output_dir / file_id
    wb.save(str(output_path))

    # ── Write audit JSON ─────────────────────────────────────────────────
    total_prior = sum(r["prior_amount_idr"] or 0.0 for r in assets)
    total_current = sum(r["current_amount_idr"] or 0.0 for r in assets)
    total_market = sum(r["market_value_idr"] or 0.0 for r in assets)
    liability_total = sum(r["current_amount_idr"] or 0.0 for r in liabilities)

    result = ExportResult(
        tax_year=tax_year,
        file_id=file_id,
        total_rows=len(assets),
        total_prior=round(total_prior, 2),
        total_current=round(total_current, 2),
        total_market=round(total_market, 2),
        liability_total=round(liability_total, 2),
        rows=audit_rows,
    )

    audit_path = output_path.with_suffix(".audit.json")
    audit_path.write_text(
        json.dumps(asdict(result), indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    return result


def _write_data_row(ws, xlsx_row: int, row: dict, audit_rows: list) -> None:
    """Write a single coretax_row into the XLSX at the given row number.

    Columns:
      A = sequential number
      B = kode_harta
      C = keterangan (short description)
      D = acquisition_year
      E = prior_amount_idr
      F = current_amount_idr (formula or literal based on source)
      G = market_value_idr
      H = keterangan (long description / notes)
    """
    # A = sequential number (auto)
    ws[f"A{xlsx_row}"] = xlsx_row - FIRST_DATA_ROW + 1

    # B = kode_harta
    ws[f"B{xlsx_row}"] = row.get("kode_harta") or ""

    # C = keterangan
    ws[f"C{xlsx_row}"] = row.get("keterangan") or ""

    # D = acquisition_year
    ws[f"D{xlsx_row}"] = row.get("acquisition_year") or ""

    # E = prior_amount_idr
    prior = row.get("prior_amount_idr")
    ws[f"E{xlsx_row}"] = prior if prior is not None else ""

    # F = current_amount_idr — THE KEY RULE
    source = row.get("current_amount_source")
    current = row.get("current_amount_idr")
    f_rule = "blank"

    if source == "carried_forward":
        ws[f"F{xlsx_row}"] = f"=E{xlsx_row}"
        f_rule = "formula"
    elif current is not None:
        ws[f"F{xlsx_row}"] = current
        f_rule = "literal"
    else:
        ws[f"F{xlsx_row}"] = ""

    # G = market_value_idr
    market = row.get("market_value_idr")
    ws[f"G{xlsx_row}"] = market if market is not None else ""

    # H = long description / notes
    notes = row.get("keterangan") or ""
    internal_notes = row.get("notes_internal") or ""
    if internal_notes:
        notes = f"{notes} | {internal_notes}" if notes else internal_notes
    ws[f"H{xlsx_row}"] = notes

    audit_rows.append(ExportAuditRow(
        xlsx_row=xlsx_row,
        stable_key=row["stable_key"],
        kode_harta=row.get("kode_harta") or "",
        keterangan=row.get("keterangan") or "",
        prior_amount_idr=prior,
        current_amount_idr=current,
        market_value_idr=market,
        current_amount_source=source,
        f_cell_rule=f_rule,
    ))


def _clear_data_row(ws, xlsx_row: int) -> None:
    """Clear a data row in the template (set cells to empty)."""
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws[f"{col}{xlsx_row}"] = ""


def _find_template(template_dir: Path) -> Path | None:
    """Find the first .xlsx file in the template directory."""
    if not template_dir.exists():
        return None
    xlsx_files = sorted(template_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsx_files[0] if xlsx_files else None


def _next_output_filename(output_dir: Path, tax_year: int) -> str:
    """Generate the next versioned filename for the output XLSX."""
    version = 1
    while True:
        filename = f"CoreTax_{tax_year}_v{version}.xlsx"
        if not (output_dir / filename).exists():
            return filename
        version += 1
