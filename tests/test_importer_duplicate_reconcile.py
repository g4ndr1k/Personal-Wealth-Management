from pathlib import Path

import openpyxl
import pytest

from finance.db import open_db
from finance.importer import direct_import


class DummyResult:
    def __init__(self, merchant=None, category=None, layer=2):
        self.merchant = merchant
        self.category = category
        self.layer = layer


class DummyCategorizer:
    def __init__(self):
        self._exact = {}
        self._regex = []
        self.categories = ['Education']

    def reload_aliases(self, aliases):
        return None

    def categorize(self, raw_description, owner=None, account=None):
        if 'BINUS' in raw_description:
            return DummyResult('Binus School Simprug', 'Education', 2)
        return DummyResult(None, 'Education', 2)


HEADERS = [
    'Owner', 'Month', 'Bank', 'Statement Type',
    'Tgl. Transaksi', 'Tgl. Tercatat', 'Keterangan',
    'Currency', 'Jumlah Valuta Asing', 'Kurs (RP)',
    'Jumlah (IDR)', 'Tipe', 'Saldo (IDR)', 'Nomor Rekening/Kartu',
]


def _write_xlsx(path: Path, description: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ALL_TRANSACTIONS'
    ws.append(HEADERS)
    ws.append([
        'Gandrik', 'Mar 2026', 'BCA', 'savings',
        '2026-03-05', '2026-03-05', description,
        'IDR', '', '', 55200000, 'debit', '', '2171138631',
    ])
    wb.save(path)


def test_direct_import_reconciles_parser_evolution_duplicate_and_preserves_override(tmp_path):
    db_path = str(tmp_path / 'finance.db')
    conn = open_db(db_path)
    stale_hash = 'stale-binus-hash'
    conn.execute(
        """
        INSERT INTO transactions (
            date, amount, raw_description, merchant, category, institution, account, owner,
            notes, hash, import_date, import_file, synced_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', ?, '2026-04-10', 'bca-mar-kddft6ok.xlsx', '2026-04-10 14:26:12')
        """,
        (
            '2026-03-05',
            -55200000.0,
            'TRSF E-BANKING DB 0403/FTFVA/WS9',
            None,
            'Education',
            'BCA',
            '2171138631',
            'Gandrik',
            stale_hash,
        ),
    )
    conn.execute(
        """
        INSERT INTO category_overrides (hash, category, merchant, notes, updated_at, updated_by)
        VALUES (?, 'Education', 'Manual Binus', '', '2026-04-11 08:00:00', 'user')
        """,
        (stale_hash,),
    )
    conn.commit()
    conn.close()

    xlsx_path = tmp_path / 'all_transactions.xlsx'
    _write_xlsx(xlsx_path, 'TRSF E-BANKING DB / 0403/FTFVA/WS95031 / 71201/BINUS S SIMP')

    stats = direct_import(
        xlsx_path=str(xlsx_path),
        db_path=db_path,
        categorizer=DummyCategorizer(),
        overwrite=False,
        dry_run=False,
        import_file_label='ALL_TRANSACTIONS.xlsx',
    )

    conn = open_db(db_path)
    rows = conn.execute(
        "SELECT hash, raw_description, merchant, category, import_file FROM transactions"
    ).fetchall()
    resolved = conn.execute(
        "SELECT hash, raw_description, merchant, category FROM transactions_resolved"
    ).fetchone()
    override_hashes = conn.execute("SELECT hash FROM category_overrides").fetchall()
    conn.close()

    assert stats['added'] == 0
    assert stats['reconciled'] == 1
    assert len(rows) == 1
    assert rows[0]['hash'] == stale_hash
    assert rows[0]['raw_description'] == 'TRSF E-BANKING DB / 0403/FTFVA/WS95031 / 71201/BINUS S SIMP'
    assert rows[0]['merchant'] == 'Binus School Simprug'
    assert rows[0]['category'] == 'Education'
    assert rows[0]['import_file'] == 'ALL_TRANSACTIONS.xlsx'
    assert resolved['hash'] == stale_hash
    assert resolved['merchant'] == 'Manual Binus'
    assert resolved['category'] == 'Education'
    assert override_hashes[0]['hash'] == stale_hash


def test_direct_import_does_not_merge_distinct_same_day_same_amount_transactions(tmp_path):
    db_path = str(tmp_path / 'finance.db')
    conn = open_db(db_path)
    conn.execute(
        """
        INSERT INTO transactions (
            date, amount, raw_description, merchant, category, institution, account, owner,
            notes, hash, import_date, import_file, synced_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', ?, '2026-04-15', 'ALL_TRANSACTIONS.xlsx', '2026-04-15 23:21:18')
        """,
        (
            '2026-03-14',
            -38999.0,
            'THE COFFEE CLUB TANGERANG KOT',
            'Dining Out',
            'Dining Out',
            'Maybank',
            '4047 76XX XXXX 6004',
            'Gandrik',
            'coffee-a',
        ),
    )
    conn.commit()
    conn.close()

    xlsx_path = tmp_path / 'coffee.xlsx'
    _write_xlsx(xlsx_path, 'THE COFFEE CLUB TANGERANG ID')

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['ALL_TRANSACTIONS']
    ws['C2'] = 'Maybank'
    ws['D2'] = 'cc'
    ws['E2'] = '2026-03-14'
    ws['F2'] = '2026-03-14'
    ws['G2'] = 'THE COFFEE CLUB TANGERANG ID'
    ws['K2'] = 38999
    ws['L2'] = 'debit'
    ws['N2'] = '4047 76XX XXXX 6004'
    wb.save(xlsx_path)

    stats = direct_import(
        xlsx_path=str(xlsx_path),
        db_path=db_path,
        categorizer=DummyCategorizer(),
        overwrite=False,
        dry_run=False,
        import_file_label='ALL_TRANSACTIONS.xlsx',
    )

    conn = open_db(db_path)
    rows = conn.execute(
        "SELECT raw_description FROM transactions WHERE institution='Maybank' ORDER BY raw_description"
    ).fetchall()
    conn.close()

    assert stats['added'] == 1
    assert stats['reconciled'] == 0
    assert [r['raw_description'] for r in rows] == [
        'THE COFFEE CLUB TANGERANG ID',
        'THE COFFEE CLUB TANGERANG KOT',
    ]


def test_direct_import_fails_fast_on_xlsx_header_mismatch(tmp_path):
    db_path = str(tmp_path / 'finance.db')
    open_db(db_path).close()

    xlsx_path = tmp_path / 'bad_headers.xlsx'
    _write_xlsx(xlsx_path, 'TRSF E-BANKING DB / 0403/FTFVA/WS95031')

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['ALL_TRANSACTIONS']
    ws['C1'], ws['D1'] = ws['D1'].value, ws['C1'].value
    wb.save(xlsx_path)

    with pytest.raises(ValueError, match="XLSX header mismatch") as excinfo:
        direct_import(
            xlsx_path=str(xlsx_path),
            db_path=db_path,
            categorizer=DummyCategorizer(),
            overwrite=False,
            dry_run=False,
            import_file_label='ALL_TRANSACTIONS.xlsx',
        )

    message = str(excinfo.value)
    assert "col 2: expected 'Bank', got 'Statement Type'" in message
    assert "col 3: expected 'Statement Type', got 'Bank'" in message
