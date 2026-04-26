"""Tests for the persistent CoreTax SPT ledger workflow."""
from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

from finance.coretax.carry_forward import commit_staging_batch
from finance.coretax.db import (
    ASSET_CODE_SEED,
    ensure_coretax_tables,
    get_mappings,
    get_rows_for_year,
    increment_mapping_hit,
    insert_row,
    make_stable_key_cash,
    make_stable_key_investment,
    update_row,
    upsert_mapping,
    upsert_taxpayer,
)
from finance.coretax.exporter import ExportError, export_coretax_xlsx
from finance.coretax.import_parser import ImportParseError, parse_prior_year_xlsx
from finance.coretax.reconcile import run_reconcile


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def conn():
    """Fresh in-memory SQLite DB with coretax tables initialized."""
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    ensure_coretax_tables(c)
    # PWM tables required by reconcile
    c.executescript("""
        CREATE TABLE account_balances (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, institution TEXT,
            account TEXT, owner TEXT, currency TEXT, balance_idr REAL
        );
        CREATE TABLE holdings (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, asset_class TEXT,
            institution TEXT, owner TEXT, currency TEXT, asset_name TEXT,
            isin_or_code TEXT, cost_basis_idr REAL, market_value_idr REAL
        );
        CREATE TABLE liabilities (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, liability_type TEXT,
            liability_name TEXT, institution TEXT, owner TEXT, balance_idr REAL
        );
    """)
    c.commit()
    yield c
    c.close()


def _make_synth_template(path: Path, prior_year: int = 2024,
                         taxpayer_name: str = "Test Taxpayer",
                         npwp: str = "0000000000000000") -> None:
    """Create a small synthetic SPT template at `path` with 4 representative rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(prior_year + 1)
    ws["B1"] = "Nama wajib pajak"
    ws["C1"] = taxpayer_name
    ws["B2"] = "NPWP"
    ws["C2"] = npwp
    ws["B3"] = "Tahun Pajak"
    ws["C3"] = prior_year
    ws["A4"] = "No"
    ws["B4"] = "Kode Harta"
    ws["C4"] = "Keterangan"
    ws["D4"] = "Tahun Perolehan"
    ws["E4"] = prior_year
    ws["F4"] = prior_year + 1
    ws["G4"] = "Nilai saat ini"
    ws["H4"] = "Keterangan"
    # Rows 6+ data
    rows = [
        # No, kode, keterangan, acq_year, E (prior), F (carry), G (mv), H (note)
        (1, "061", "Tanah & Bangunan",  2010, 100000000, 100000000, 500000000, "Property A"),
        (2, "012", "Tabungan",          2020,  10000000,         0,         0, "BCA acct 1234567890 an Test Taxpayer"),
        (3, "039", "Saham",             2022, 200000000,         0, 250000000, "Indopremier IDD ABC123 an Test Taxpayer"),
        (4, "043", "Mobil",             2018, 300000000, 300000000, 200000000, "Test Car"),
    ]
    for i, (no, kode, ket, ay, e, f, g, h) in enumerate(rows, start=6):
        ws[f"A{i}"] = no
        ws[f"B{i}"] = kode
        ws[f"C{i}"] = ket
        ws[f"D{i}"] = ay
        ws[f"E{i}"] = e
        ws[f"F{i}"] = f
        ws[f"G{i}"] = g
        ws[f"H{i}"] = h
    # Footer (row 48+)
    ws["C48"] = "TOTAL ASET KOTOR"
    ws["E48"] = "=SUM(E6:E47)"
    ws["F48"] = "=SUM(F6:F47)"
    ws["C49"] = "KENAIKAN ASET KOTOR"
    ws["F49"] = "=F48-E48"
    wb.save(str(path))


@pytest.fixture
def synth_template(tmp_path):
    p = tmp_path / "CoreTax_2024.xlsx"
    _make_synth_template(p, prior_year=2024)
    return p


# ── Tests ────────────────────────────────────────────────────────────────────

def test_import_parser_synthetic_template(conn, synth_template):
    """Parser successfully reads a well-formed prior-year template."""
    result = parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    assert result["row_count"] == 4
    assert result["prior_tax_year"] == 2024
    assert result["warnings"] == []
    # Each staged row must carry raw col E/F/G/H values for audit
    rows = conn.execute(
        "SELECT source_row_no, source_col_e_value, source_col_f_value, "
        "source_col_g_value, source_col_h_note FROM coretax_import_staging"
    ).fetchall()
    assert len(rows) == 4
    for r in rows:
        assert r["source_row_no"] >= 6
        assert r["source_col_e_value"]  # non-empty
        assert r["source_col_h_note"]   # non-empty


def test_target_year_mismatch_rejected(conn, synth_template):
    """G5: Parser rejects a template whose F header doesn't equal target_tax_year."""
    with pytest.raises(ImportParseError) as exc_info:
        parse_prior_year_xlsx(synth_template, target_tax_year=2026, conn=conn)
    assert "year mismatch" in str(exc_info.value).lower()
    assert "2024" in str(exc_info.value)


def test_f_header_mismatch_rejected(conn, synth_template):
    """G5: Parser rejects a workbook whose F header is not target_tax_year."""
    wb = openpyxl.load_workbook(str(synth_template))
    ws = wb.worksheets[0]
    ws["F4"] = 2026
    wb.save(str(synth_template))

    with pytest.raises(ImportParseError) as exc_info:
        parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    assert "year mismatch" in str(exc_info.value).lower()
    assert "F header=2026" in str(exc_info.value)


def test_carry_forward_commit_splits_correctly(conn, synth_template):
    """Sticky codes carry forward; refreshable codes stay unset."""
    result = parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    commit_staging_batch(conn, result["batch_id"], 2025)

    rows = get_rows_for_year(conn, 2025)
    assert len(rows) == 4
    by_kode = {r["kode_harta"]: r for r in rows}

    # 061 (Tanah & Bangunan) — sticky
    assert by_kode["061"]["current_amount_source"] == "carried_forward"
    assert by_kode["061"]["current_amount_idr"] == 100000000

    # 043 (Mobil) — sticky
    assert by_kode["043"]["current_amount_source"] == "carried_forward"
    assert by_kode["043"]["current_amount_idr"] == 300000000

    # 012 (Tabungan) — refreshable, must NOT be pre-zeroed
    assert by_kode["012"]["current_amount_source"] == "unset"
    assert by_kode["012"]["current_amount_idr"] is None

    # 039 (Saham) — refreshable
    assert by_kode["039"]["current_amount_source"] == "unset"
    assert by_kode["039"]["current_amount_idr"] is None


def test_lock_blocks_auto_reconcile(conn):
    """G2 part 1: amount_locked=1 prevents reconcile from overwriting current_amount_idr."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "12345")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA acct 12345",
        institution="BCA", account_number_masked="12345",
        current_amount_idr=999999.0, current_amount_source="manual",
        amount_locked=1, locked_reason="manual edit",
    )
    # Seed PWM that would otherwise overwrite
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "12345", "Test", "IDR", 50000.0),
    )
    # Seed mapping pointing at this row
    upsert_mapping(conn, "account_number", "12345", "012", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    result = run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 999999.0  # unchanged
    # Trace must record skipped_locked
    statuses = [t["status"] for t in result["trace"]]
    assert "locked_skipped" in statuses


def test_market_value_lock_independent_of_amount_lock(conn):
    """G2 part 2: market_value_locked alone blocks only market_value_idr writes."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_investment("stock", "Indopremier", "ABC123", "Test")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="039", keterangan="Indopremier ABC123",
        institution="Indopremier", external_ref="ABC123",
        market_value_idr=111111.0, market_value_source="manual",
        market_value_locked=1, amount_locked=0,
    )
    conn.execute(
        "INSERT INTO holdings (snapshot_date, asset_class, institution, owner, currency, "
        "asset_name, isin_or_code, cost_basis_idr, market_value_idr) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "stock", "Indopremier", "Test", "IDR", "ABC", "ABC123", 200000.0, 250000.0),
    )
    upsert_mapping(conn, "isin", "abc123", "039", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 200000.0          # written (unlocked)
    assert after["market_value_idr"] == 111111.0            # untouched (locked)


def test_cash_reconcile_does_not_write_market_value(conn):
    """G2 part 3: cash kode 012 must never write market_value, even when unlocked."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "999")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA 999",
        institution="BCA", account_number_masked="999",
        market_value_idr=None, market_value_source="unset",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "999", "Test", "IDR", 75000.0),
    )
    upsert_mapping(conn, "account_number", "999", "012", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 75000.0   # cash amount written
    assert after["market_value_idr"] is None        # market value untouched


def test_mapping_hit_increments_counter(conn):
    """G1: mapping hits increment and last_used_tax_year updates on successful apply."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "11111")
    insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA 11111",
        institution="BCA",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "11111", "Test", "IDR", 1000.0),
    )
    mid = upsert_mapping(conn, "account_number", "11111", "012", "asset",
                         target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 1
    assert m["last_used_tax_year"] == 2025

    # Re-run; hits must increment again
    run_reconcile(conn, 2025, "2025-01", "2025-12")
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 2


def test_create_from_unmatched_persists_mapping(conn):
    """G1 PWA round-trip: simulate the create-from-unmatched flow.

    1) Reconcile produces an unmatched PWM row with proposed_match_kind/value.
    2) UI calls /api/coretax/rows to create a manual ledger row.
    3) UI calls /api/coretax/mappings to persist the learned mapping.
    4) Re-run reconcile — same PWM source now auto-applies; hits increments.
    """
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "Permata", "9876", "Test", "IDR", 5000.0),
    )
    conn.commit()

    # Step 1: initial reconcile produces an unmatched row.
    r1 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert r1["summary"]["unmatched"] == 1
    um = r1["unmatched"][0]
    assert um["payload"]["proposed_match_kind"] == "account_number"
    assert um["payload"]["proposed_match_value"] == "9876"

    # Step 2 + 3: UI creates a row from the unmatched payload and persists mapping.
    sk = make_stable_key_cash("Permata", "9876")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="Permata 9876",
        institution="Permata",
    )
    mid = upsert_mapping(conn, "account_number", "9876", "012", "asset",
                         target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()
    assert any(m["id"] == mid for m in get_mappings(conn))

    # Step 4: re-run — should apply, not appear as unmatched.
    r2 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert r2["summary"]["unmatched"] == 0
    assert r2["summary"]["filled"] == 1
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 5000.0
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 1


def test_export_capacity_guard(conn, tmp_path):
    """G plan #10: more than 42 asset rows must fail with no XLSX written."""
    # Need a template to load — build a minimal one
    template_dir = tmp_path / "templates"
    template_dir.mkdir()
    output_dir = tmp_path / "out"
    template_path = template_dir / "CoreTax_template.xlsx"
    _make_synth_template(template_path, prior_year=2024)

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    for i in range(43):
        insert_row(
            conn, tax_year=2025, kind="asset",
            stable_key=f"manual:061:row-{i}:2020:abc{i:04d}",
            kode_harta="061", keterangan=f"Property {i}",
            current_amount_idr=1000.0,
        )
    conn.commit()

    with pytest.raises(ExportError) as exc_info:
        export_coretax_xlsx(conn, 2025, template_dir, output_dir)
    assert "capacity" in str(exc_info.value).lower()
    # No partial XLSX must exist
    assert not any(output_dir.glob("*.xlsx")) if output_dir.exists() else True


def test_export_f_cell_formula_rule(conn, tmp_path):
    """Plan #9: F cell is `=E{n}` iff current_amount_source='carried_forward'.

    Do NOT infer formula-vs-literal from numeric equality between E and F.
    """
    template_dir = tmp_path / "templates"
    template_dir.mkdir()
    output_dir = tmp_path / "out"
    template_path = template_dir / "CoreTax_template.xlsx"
    _make_synth_template(template_path, prior_year=2024)

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    # Row A: carried forward — must produce '=E6'
    insert_row(
        conn, tax_year=2025, kind="asset",
        stable_key="manual:061:carry:2020:aaaa1111",
        kode_harta="061", keterangan="Carry Property",
        prior_amount_idr=500.0, current_amount_idr=500.0,
        prior_amount_source="imported",
        current_amount_source="carried_forward",
    )
    # Row B: literal — value coincidentally equals prior, but source is 'manual'
    insert_row(
        conn, tax_year=2025, kind="asset",
        stable_key="manual:039:literal:2022:bbbb2222",
        kode_harta="039", keterangan="Literal Stock",
        prior_amount_idr=700.0, current_amount_idr=700.0,
        prior_amount_source="imported",
        current_amount_source="manual",
    )
    conn.commit()

    result = export_coretax_xlsx(conn, 2025, template_dir, output_dir)
    out_path = output_dir / result.file_id
    wb = openpyxl.load_workbook(str(out_path), data_only=False)
    ws = wb.worksheets[0]

    # Rows are sorted by kode_harta then id: 039 first (id=2 doesn't matter,
    # 039 < 061 alphabetically). Find them deterministically by content.
    # Row 6 = first asset, row 7 = second asset.
    cells = {6: ws["F6"].value, 7: ws["F7"].value}
    ket_to_f = {ws[f"C{r}"].value: f for r, f in cells.items()}

    # Carry row → formula
    assert ket_to_f["Carry Property"] == "=E7" or ket_to_f["Carry Property"] == "=E6"
    assert isinstance(ket_to_f["Carry Property"], str)
    assert ket_to_f["Carry Property"].startswith("=E")
    # Literal row → numeric, NOT formula
    assert ket_to_f["Literal Stock"] == 700.0
    assert not (isinstance(ket_to_f["Literal Stock"], str) and
                str(ket_to_f["Literal Stock"]).startswith("=E"))


def test_check_constraint_rejects_typo(conn):
    """Plan #15: CHECK constraint catches enum typos."""
    with pytest.raises(sqlite3.IntegrityError):
        conn.execute(
            "INSERT INTO coretax_rows (tax_year, kind, stable_key, current_amount_source, "
            "created_at, updated_at) VALUES (2025, 'asset', 'k1', 'auto_reconcile', 'now', 'now')"
        )
