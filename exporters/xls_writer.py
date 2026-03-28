"""
XLS Exporter — redesigned for multi-owner, multi-bank output.

File layout:
  output/xls/
    {Bank}_{Owner}.xlsx          e.g. Maybank_Gandrik.xlsx, BCA_Helen.xlsx
    ALL_TRANSACTIONS.xlsx        flat table, all banks + owners, Owner column

Sheet naming inside per-person-per-bank files:
  "{Mon YYYY} CC"      credit card statement
  "{Mon YYYY} Savings" savings / tabungan statement
  "{Mon YYYY} Consol"  consolidated statement

Sheet naming map:
  statement_type → suffix
  "cc"           → "CC"
  "savings"      → "Savings"
  "consolidated" → "Consol"

ALL_TRANSACTIONS.xlsx:
  Single sheet "ALL_TRANSACTIONS"
  Columns: Owner | Month | Bank | Statement Type | Tgl. Transaksi | Tgl. Tercatat |
           Keterangan | Currency | Jumlah Valuta Asing | Kurs (RP) |
           Jumlah (IDR) | Tipe | Saldo (IDR) | Nomor Rekening/Kartu

Color coding (industry standard):
  Blue text  = hardcoded input values from statement
  Black text = formula cells
  Yellow bg  = key summary figures
"""
import os
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from parsers.base import StatementResult, Transaction, AccountSummary
from parsers.owner import detect_owner

# ── Style constants ────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
BLUE       = Font(name=FONT_NAME, color="0000FF", size=10)
BLACK      = Font(name=FONT_NAME, color="000000", size=10)
WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=10)
HEADER_FONT= Font(name=FONT_NAME, bold=True, size=10)

FILL_HEADER  = PatternFill("solid", start_color="2F4F8F")
FILL_SECTION = PatternFill("solid", start_color="FFC000")
FILL_YELLOW  = PatternFill("solid", start_color="FFFF00")
FILL_CREDIT  = PatternFill("solid", start_color="E8F5E9")
FILL_DEBIT   = PatternFill("solid", start_color="FFF3E0")

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
IDR_FMT = '#,##0;(#,##0);"-"'

# Statement type → sheet suffix
_TYPE_SUFFIX = {
    "cc":           "CC",
    "savings":      "Savings",
    "consolidated": "Consol",
}

MONTHS_ID = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}

ALL_TX_HEADERS = [
    "Owner", "Month", "Bank", "Statement Type",
    "Tgl. Transaksi", "Tgl. Tercatat", "Keterangan",
    "Currency", "Jumlah Valuta Asing", "Kurs (RP)",
    "Jumlah (IDR)", "Tipe", "Saldo (IDR)", "Nomor Rekening/Kartu",
]
ALL_TX_WIDTHS = [12, 12, 10, 16, 14, 14, 42, 10, 18, 14, 18, 10, 18, 22]


# ── Public entry point ─────────────────────────────────────────────────────────
def export(result: StatementResult, output_dir: str,
           owner_mappings: dict = None) -> tuple[str, str]:
    """
    Write per-person-per-bank file and update ALL_TRANSACTIONS.xlsx.
    Returns (per_person_path, all_tx_path).
    """
    os.makedirs(output_dir, exist_ok=True)

    owner = result.owner if result.owner else detect_owner(result.customer_name, owner_mappings)
    result.owner = owner  # attach for use in writers

    # ── Per-person-per-bank file ──────────────────────────────────────────
    filename = f"{result.bank}_{owner}.xlsx"
    filepath = os.path.join(output_dir, filename)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = result.sheet_name if result.sheet_name else _sheet_name(result)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    _sort_sheets(wb)

    if result.statement_type == "cc":
        _write_cc_sheet(ws, result)
    elif result.statement_type == "savings":
        _write_savings_sheet(ws, result)
    else:
        _write_consol_sheet(ws, result)

    wb.save(filepath)

    # ── ALL_TRANSACTIONS.xlsx ─────────────────────────────────────────────
    all_tx_path = _update_all_transactions(output_dir, result, sheet_name, owner)

    return filepath, all_tx_path


# ── Sheet name helpers ─────────────────────────────────────────────────────────
def _sheet_name(result: StatementResult) -> str:
    """e.g. 'Mar 2026 CC', 'Feb 2026 Savings', 'Feb 2026 Consol'"""
    date_str = result.print_date or result.period_end
    month_label = _month_label(date_str)
    suffix = _TYPE_SUFFIX.get(result.statement_type, result.statement_type.title())
    return f"{month_label} {suffix}"


def _month_label(date_str: str) -> str:
    """DD/MM/YYYY → 'Feb 2026'"""
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", date_str)
    if m:
        _, month, year = m.groups()
        return f"{MONTHS_ID.get(month, month)} {year}"
    now = datetime.now()
    return f"{MONTHS_ID[str(now.month).zfill(2)]} {now.year}"


def _sort_sheets(wb: Workbook):
    def _key(name):
        m = re.match(r"(\w+)\s+(\d{4})", name)
        if m:
            mon, yr = m.groups()
            mon_order = list(MONTHS_ID.values()).index(mon) + 1 if mon in MONTHS_ID.values() else 0
            return (int(yr), mon_order, name)
        return (9999, 99, name)
    wb._sheets.sort(key=lambda ws: _key(ws.title))


# ── CC sheet ───────────────────────────────────────────────────────────────────
def _write_cc_sheet(ws, result: StatementResult):
    row = _write_meta_block(ws, 1, result)

    if result.accounts:
        acc = result.accounts[0]
        row = _section_header(ws, row, "RINGKASAN TAGIHAN KARTU KREDIT")
        row = _table_header(ws, row, ["Field", "Value"], [32, 22])
        for label, val in [
            ("Nomor Kartu",           acc.account_number or ""),
            ("Total Tagihan (IDR)",   _fmt(acc.closing_balance)),
            ("Pembayaran Minimum (IDR)", _fmt(acc.extra.get("min_payment"))),
            ("Limit Gabungan (IDR)",  _fmt(acc.credit_limit)),
            ("Tanggal Jatuh Tempo",   acc.extra.get("due_date", "")),
            ("Tagihan Sebelumnya (IDR)", _fmt(acc.extra.get("prev_balance"))),
            ("Total Pembelanjaan (IDR)", _fmt(acc.extra.get("purchases"))),
            ("Total Pembayaran (IDR)", _fmt(acc.extra.get("payments"))),
        ]:
            ws.cell(row, 1, label).font = BLACK
            ws.cell(row, 1).alignment = LEFT
            c = ws.cell(row, 2, val)
            c.font = BLUE; c.alignment = RIGHT
            _border(ws, row, 1, 2)
            row += 1
        row += 1

    row = _section_header(ws, row, "DETAIL TRANSAKSI")
    hdrs = ["Tgl. Transaksi","Tgl. Tercatat","Keterangan",
            "Currency","Jumlah Valuta Asing","Kurs (RP)","Jumlah (IDR)","Tipe"]
    row = _table_header(ws, row, hdrs, [14,14,42,10,18,14,18,10])

    for tx in result.transactions:
        fill = FILL_CREDIT if tx.tx_type == "Credit" else FILL_DEBIT
        for col, val in enumerate([
            tx.date_transaction, tx.date_posted or "",
            tx.description, tx.currency,
            tx.foreign_amount or "", tx.exchange_rate or "",
            tx.amount_idr, tx.tx_type,
        ], 1):
            c = ws.cell(row, col, val)
            c.font = BLUE; c.fill = fill
            c.alignment = LEFT if col == 3 else RIGHT
            _border(ws, row, col, col)
            if col in (5, 6, 7) and val != "":
                c.number_format = IDR_FMT
        row += 1

    total_debit  = sum(t.amount_idr for t in result.transactions if t.tx_type == "Debit")
    total_credit = sum(t.amount_idr for t in result.transactions if t.tx_type == "Credit")
    for label, val in [("TOTAL DEBIT", total_debit), ("TOTAL KREDIT", total_credit)]:
        ws.cell(row, 6, label).font = HEADER_FONT
        c = ws.cell(row, 7, val)
        c.font = HEADER_FONT; c.number_format = IDR_FMT; c.fill = FILL_YELLOW
        row += 1


# ── Savings sheet ──────────────────────────────────────────────────────────────
def _write_savings_sheet(ws, result: StatementResult):
    row = _write_meta_block(ws, 1, result)

    if result.accounts:
        acc = result.accounts[0]
        row = _section_header(ws, row, "RINGKASAN REKENING")
        row = _table_header(ws, row, ["Field", "Value"], [32, 22])
        for label, val in [
            ("Nomor Rekening",    acc.account_number or ""),
            ("Mata Uang",         acc.currency),
            ("Saldo Awal (IDR)",  _fmt(acc.opening_balance)),
            ("Saldo Akhir (IDR)", _fmt(acc.closing_balance)),
            ("Total Kredit (IDR)",_fmt(acc.total_credit)),
            ("Total Debit (IDR)", _fmt(acc.total_debit)),
            ("Periode",           acc.extra.get("period", "")),
        ]:
            ws.cell(row, 1, label).font = BLACK; ws.cell(row, 1).alignment = LEFT
            c = ws.cell(row, 2, val); c.font = BLUE; c.alignment = RIGHT
            _border(ws, row, 1, 2); row += 1
        row += 1

    row = _section_header(ws, row, "MUTASI TRANSAKSI")
    hdrs = ["Tgl. Transaksi","Keterangan","Mutasi Debet","Mutasi Kredit","Saldo","Tipe"]
    row = _table_header(ws, row, hdrs, [14,42,18,18,18,10])

    for tx in result.transactions:
        fill = FILL_CREDIT if tx.tx_type == "Credit" else FILL_DEBIT
        debit_val  = tx.amount_idr if tx.tx_type == "Debit"   else ""
        credit_val = tx.amount_idr if tx.tx_type == "Credit"  else ""
        for col, val in enumerate([
            tx.date_transaction, tx.description,
            debit_val, credit_val,
            tx.balance if tx.balance is not None else "",
            tx.tx_type,
        ], 1):
            c = ws.cell(row, col, val)
            c.font = BLUE; c.fill = fill
            c.alignment = LEFT if col == 2 else RIGHT
            _border(ws, row, col, col)
            if col in (3, 4, 5) and val != "":
                c.number_format = IDR_FMT
        row += 1


# ── Consolidated sheet ─────────────────────────────────────────────────────────
def _write_consol_sheet(ws, result: StatementResult):
    row = _write_meta_block(ws, 1, result)

    row = _section_header(ws, row, "RINGKASAN ALOKASI ASET & PINJAMAN")
    row = _table_header(ws, row, ["Kategori / Produk","Mata Uang","Saldo","Catatan"], [42,12,22,30])
    for acc in result.accounts:
        ws.cell(row,1,acc.product_name).font = BLUE
        ws.cell(row,1).alignment = LEFT
        ws.cell(row,2,acc.currency).font = BLUE
        ws.cell(row,2).alignment = CENTER
        c = ws.cell(row,3,acc.closing_balance or 0)
        c.font = BLUE; c.alignment = RIGHT; c.number_format = IDR_FMT
        notes = "; ".join(f"{k}={v}" for k,v in acc.extra.items() if v is not None)
        ws.cell(row,4,notes).font = BLACK
        _border(ws, row, 1, 4); row += 1
    row += 1

    if result.transactions:
        row = _section_header(ws, row, "MUTASI TRANSAKSI TABUNGAN")
        hdrs = ["Tgl. Transaksi","Keterangan","Mutasi Debet","Mutasi Kredit","Saldo","Mata Uang","Tipe"]
        row = _table_header(ws, row, hdrs, [14,42,18,18,18,12,10])
        for tx in result.transactions:
            fill = FILL_CREDIT if tx.tx_type == "Credit" else FILL_DEBIT
            for col, val in enumerate([
                tx.date_transaction, tx.description,
                tx.amount_idr if tx.tx_type == "Debit"  else "",
                tx.amount_idr if tx.tx_type == "Credit" else "",
                tx.balance if tx.balance is not None else "",
                tx.currency,
                tx.tx_type,
            ], 1):
                c = ws.cell(row, col, val)
                c.font = BLUE; c.fill = fill
                c.alignment = LEFT if col == 2 else RIGHT
                _border(ws, row, col, col)
                if col in (3, 4, 5) and val != "":
                    c.number_format = IDR_FMT
            row += 1
        row += 1

    if result.exchange_rates:
        row = _section_header(ws, row, "NILAI TUKAR RUPIAH")
        row = _table_header(ws, row, ["Mata Uang","Kurs (IDR)"], [14,18])
        for ccy, rate in result.exchange_rates.items():
            ws.cell(row,1,ccy).font = BLUE; ws.cell(row,1).alignment = CENTER
            c = ws.cell(row,2,rate); c.font = BLUE; c.alignment = RIGHT
            c.number_format = IDR_FMT; _border(ws, row, 1, 2); row += 1


# ── ALL_TRANSACTIONS.xlsx ──────────────────────────────────────────────────────
def _update_all_transactions(output_dir: str, result: StatementResult,
                             month_label_with_type: str, owner: str) -> str:
    """
    Maintain a single ALL_TRANSACTIONS.xlsx in output_dir.
    Removes existing rows matching (owner, month_label, bank, stmt_type,
    account_numbers), then appends new rows.

    Account numbers are included in the dedup key so that multiple accounts
    from the same bank/owner/month (e.g. two Permata savings accounts for
    the same person) don't overwrite each other.
    """
    filepath = os.path.join(output_dir, "ALL_TRANSACTIONS.xlsx")
    month_label = " ".join(month_label_with_type.split()[:2])  # "Feb 2026" from "Feb 2026 CC"

    # Collect all account numbers in this result for scoped dedup
    result_accounts = {tx.account_number for tx in result.transactions if tx.account_number}

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        # Remove rows matching this exact owner/month/bank/type/account combo
        rows_to_delete = [
            r for r in range(2, ws.max_row + 1)
            if (ws.cell(r, 1).value == owner and
                ws.cell(r, 2).value == month_label and
                ws.cell(r, 3).value == result.bank and
                ws.cell(r, 4).value == result.statement_type and
                (not result_accounts or ws.cell(r, 14).value in result_accounts))
        ]
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ALL_TRANSACTIONS"
        _table_header(ws, 1, ALL_TX_HEADERS, ALL_TX_WIDTHS)
        start_row = 2

    for tx in result.transactions:
        row_data = [
            owner,
            month_label,
            result.bank,
            result.statement_type,
            tx.date_transaction,
            tx.date_posted or "",
            tx.description,
            tx.currency,
            tx.foreign_amount or "",
            tx.exchange_rate or "",
            tx.amount_idr,
            tx.tx_type,
            tx.balance if tx.balance is not None else "",
            tx.account_number,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(start_row, col, val)
            c.font = BLUE
            c.alignment = LEFT if col == 7 else RIGHT
            _border(ws, start_row, col, col)
            if col in (9, 10, 11, 13) and val != "":
                c.number_format = IDR_FMT
        start_row += 1

    wb.save(filepath)
    return filepath


# ── Shared formatting helpers ──────────────────────────────────────────────────
def _write_meta_block(ws, row: int, result: StatementResult) -> int:
    owner = getattr(result, "owner", "Unknown")
    stmt_label = {"cc": "Tagihan Kartu Kredit", "savings": "Rekening Tabungan",
                  "consolidated": "Consolidated Statement"}.get(result.statement_type, result.statement_type)
    for label, val in [
        ("Bank",          result.bank),
        ("Owner",         owner),
        ("Tipe Laporan",  stmt_label),
        ("Nama Nasabah",  result.customer_name),
        ("Periode",       f"{result.period_start} – {result.period_end}"),
        ("Tanggal Cetak", result.print_date or ""),
    ]:
        ws.cell(row, 1, label).font = HEADER_FONT
        ws.cell(row, 1).alignment = LEFT
        ws.cell(row, 2, val).font = BLUE
        ws.cell(row, 2).alignment = LEFT
        row += 1
    return row + 1


def _section_header(ws, row: int, title: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row, 1, title)
    c.font = WHITE_BOLD; c.fill = FILL_SECTION; c.alignment = CENTER
    return row + 1


def _table_header(ws, row: int, headers: list, col_widths: list = None) -> int:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
        if col_widths and col - 1 < len(col_widths):
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]
    return row + 1


def _border(ws, row: int, col_start: int, col_end: int):
    for col in range(col_start, col_end + 1):
        ws.cell(row, col).border = BORDER


def _fmt(val) -> str:
    if val is None: return ""
    try: return f"{float(val):,.0f}"
    except: return str(val)
